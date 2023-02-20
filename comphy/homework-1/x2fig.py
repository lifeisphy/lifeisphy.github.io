'''
辅助作图所使用的程序
'''
import matplotlib.pyplot as plt 
import numpy as np
from numpy.lib.function_base import average
from scipy import stats
from math import log10
def get_linear_graph_by_xls(filename,xrange,yrange,xlty=True,tofile="test.png"):
    '''
    xlty：x的标号是否小于y
    '''
    if xrange[1]-xrange[0]==1:
        title_in_row=True
    elif yrange[1]-yrange[0]==1:
        title_in_row=False
    else:
        print('error')
        exit(-1)

    from openpyxl import load_workbook
    wb=load_workbook(filename,data_only=True)
    ws=wb['Sheet1']
    x=[]
    y=[]
    xlabel,ylabel='',''
    if title_in_row:
        if xlty:#左列x,右列y
            xlabel=ws.cell(yrange[0],xrange[0]).value
            ylabel=ws.cell(yrange[0],xrange[1]).value
            for row in range(yrange[0]+1,yrange[1]+1):
                x.append(ws.cell(row,xrange[0]).value)
                y.append(ws.cell(row,xrange[1]).value)
        else:#左列y,右列x
            xlabel=ws.cell(yrange[0],xrange[1]).value
            ylabel=ws.cell(yrange[0],xrange[0]).value
            for row in range(yrange[0]+1,yrange[1]+1):
                x.append(ws.cell(row,xrange[1]).value)
                y.append(ws.cell(row,xrange[0]).value)
    else:
        if xlty:#上列x,下列y
            xlabel=ws.cell(yrange[0],xrange[0]).value
            ylabel=ws.cell(yrange[1],xrange[0]).value
            for col in range(xrange[0]+1,xrange[1]+1):
                x.append(ws.cell(yrange[0],col).value)
                y.append(ws.cell(yrange[1],col).value)
        else:#上列y,下列x
            xlabel=ws.cell(yrange[1],xrange[0]).value
            ylabel=ws.cell(yrange[0],xrange[0]).value
            for col in range(xrange[0]+1,xrange[1]+1):
                x.append(ws.cell(yrange[1],col).value)
                y.append(ws.cell(yrange[0],col).value)
    print(x,y,xlabel,ylabel)
    get_linear_graph(x,y,xlabel,ylabel,filename=tofile)
def get_graph_by_xls(filename,xrange,yrange,xlty=True,tofile="test.png"):
    '''
    xlty：x的标号是否小于y
    '''
    if xrange[1]-xrange[0]==1:
        title_in_row=True
    elif yrange[1]-yrange[0]==1:
        title_in_row=False
    else:
        print('error')
        exit(-1)

    from openpyxl import load_workbook
    wb=load_workbook(filename,data_only=True)
    ws=wb['Sheet1']
    x=[]
    y=[]
    xlabel,ylabel='',''
    if title_in_row:
        if xlty:#左列x,右列y
            xlabel=ws.cell(yrange[0],xrange[0]).value
            ylabel=ws.cell(yrange[0],xrange[1]).value
            for row in range(yrange[0]+1,yrange[1]+1):
                x.append(ws.cell(row,xrange[0]).value)
                y.append(ws.cell(row,xrange[1]).value)
        else:#左列y,右列x
            xlabel=ws.cell(yrange[0],xrange[1]).value
            ylabel=ws.cell(yrange[0],xrange[0]).value
            for row in range(yrange[0]+1,yrange[1]+1):
                x.append(ws.cell(row,xrange[1]).value)
                y.append(ws.cell(row,xrange[0]).value)
    else:
        if xlty:#上列x,下列y
            xlabel=ws.cell(yrange[0],xrange[0]).value
            ylabel=ws.cell(yrange[1],xrange[0]).value
            for col in range(xrange[0]+1,xrange[1]+1):
                x.append(ws.cell(yrange[0],col).value)
                y.append(ws.cell(yrange[1],col).value)
        else:#上列y,下列x
            xlabel=ws.cell(yrange[1],xrange[0]).value
            ylabel=ws.cell(yrange[0],xrange[0]).value
            for col in range(xrange[0]+1,xrange[1]+1):
                x.append(ws.cell(yrange[1],col).value)
                y.append(ws.cell(yrange[0],col).value)

    #---
    x=list(map(float,x))
    y=list(map(float,y))
    x=list(map(lambda x:round(x,1),x))
    #---
    print(x,y,xlabel,ylabel)
    get_graph(x,y,xlabel,ylabel,filename=tofile)
def get_linear_graph(x,y,xlabel='',ylabel='',xname='x',yname='y',filename='test.png'):
    x_s=np.linspace(min(x),max(x),endpoint=True)
    k,b,rvalue,pvalue,stderr=stats.linregress(x,y)
    print(k,b,rvalue,pvalue,stderr)
    # linear_model=np.polyfit(T,Uout,1) #给出n次拟合的n+1个系数，降次排列
    linear_model_fn=np.poly1d([k,b])
    plt.scatter(x,y)
    plt.plot(x_s,linear_model_fn(x_s),color="black")
    plt.xlabel(xlabel)
    plt.ylabel(ylabel)
    plt.text(min(x),(min(y)+max(y))/2,"{}={:.4f}{}+{:.4f}\nr={:.6f}".format(yname,k,xname,b,rvalue))
    # plt.text(0.0,60,'Uout={:.4f}T+{:.4f}\nr={:.6f}'.format(*linear_model_fn,rvalue))
    plt.savefig(filename)
def get_uncertainty(x,y):
    '''
    计算最小二乘法中截距和斜率的不确定度
    '''
    n=len(x)
    avg_x=average(x)
    avg_y=average(y)
    avg_x_sqr=0
    avg_y_sqr=0
    avg_xy=0
    for i in range(n):
        avg_x_sqr+=x[i]**2
        avg_xy+=x[i]*y[i]
        avg_y_sqr+=y[i]**2
    avg_x_sqr/=n
    avg_xy/=n
    avg_y_sqr/=n

    k=(avg_xy-avg_x*avg_y)/(avg_x_sqr-avg_x**2)
    b=avg_y-k*avg_x
    
    sigma=0 #剩余方差
    for i in range(n):
        sigma+=(y[i]-b-k*x[i])**2
    sigma=(sigma/(n-2))**(1/2)
    r=(avg_xy-avg_x*avg_y)/(avg_x_sqr-avg_x**2)**0.5/(avg_y_sqr-avg_y**2)**0.5

    sigma_k=sigma/(n*(avg_x_sqr-avg_x**2))**0.5
    sigma_b=sigma_k*(avg_x_sqr)**0.5
    print("斜率：{}\n截距：{}\n相关系数：{}\n剩余方差：{}\nsigma_k:{}\nsigma_b:{}\n".format(k,b,r,sigma,sigma_k,sigma_b))
    return sigma_k,sigma_b
def get_graph(x,y,xlabel='',ylabel='',filename='graph.png'):
    plt.scatter(x,y,s=3)
    plt.xlabel(xlabel)
    plt.ylabel(ylabel)
    plt.show()
    plt.savefig(filename)
    plt.cla()
    
